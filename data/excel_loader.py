"""
엑셀 파일에서 데이터를 로드하는 모듈
- KPI 값, CASE 매트릭스, 이슈카드를 엑셀에서 직접 읽음
- 하드코딩 없음, 모든 데이터는 엑셀 기반
"""
import openpyxl
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from data.data_models import KPISet, Case, IssueCard


class ExcelDataLoader:
    """엑셀 데이터 로더"""

    def __init__(self, excel_path: str = "BIM 효과 정량화를 위한 멀티 에이전트 시뮬레이션 연구.xlsx"):
        """
        Args:
            excel_path: 엑셀 파일 경로
        """
        self.excel_path = excel_path
        self.workbook = None
        self._load_workbook()

    def _load_workbook(self):
        """워크북 로드"""
        if not Path(self.excel_path).exists():
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {self.excel_path}")
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)

    def load_kpi_values(self, grade: str) -> Tuple[Dict[str, float], Dict[str, float]]:
        """
        KPI 값 로드 (A/B/C/D 등급별)

        Args:
            grade: A, B, C, D

        Returns:
            (bim_kpi_dict, trad_kpi_dict) 튜플
        """
        ws = self.workbook['BIM 멀티에이전트 핵심지표_유영민']

        # 등급 컬럼 찾기 (Row 4에 A, B, C, D가 있음)
        grade_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(4, col_idx).value == grade:
                grade_col = col_idx
                break

        if grade_col is None:
            raise ValueError(f"등급 '{grade}'를 찾을 수 없습니다")

        # BIM KPI (Row 5-8: WD, CD, AF, PL)
        bim_kpi = {
            "WD": self._safe_float(ws.cell(5, grade_col).value) / 100.0,   # Warning Density
            "CD": self._safe_float(ws.cell(6, grade_col).value) / 100.0,   # Clash Density
            "AF": self._safe_float(ws.cell(7, grade_col).value) / 100.0,   # Attribute Fill
            "PL": self._safe_float(ws.cell(8, grade_col).value) / 100.0,   # Phase Link
        }

        # 전통 방식 KPI (Row 9-12: RR, SR, CR, FC)
        trad_kpi = {
            "RR": self._safe_float(ws.cell(9, grade_col).value) / 100.0,    # Rework Ratio
            "SR": self._safe_float(ws.cell(10, grade_col).value) / 100.0,   # Safety Risk
            "CR": self._safe_float(ws.cell(11, grade_col).value) / 100.0,   # Construction Delay Ratio
            "FC": (self._safe_float(ws.cell(12, grade_col).value) / 100.0) if ws.cell(12, grade_col).value else 0.0,   # Frequency of Change
        }

        return bim_kpi, trad_kpi

    def load_case_from_matrix(self,
                               location_type: str,
                               zoning: str,
                               detailed_zoning: str,
                               far: float) -> Case:
        """
        CASE 매트릭스에서 케이스 정보 로드

        Args:
            location_type: 도심/외곽
            zoning: 상업지역, 주거지역 등
            detailed_zoning: 중심상업지역, 제1종 일반주거지역 등
            far: 용적률 (%)

        Returns:
            Case 객체
        """
        # CASE 시트에서 등급 결정
        grade = self._determine_grade_from_excel(location_type, zoning, detailed_zoning, far)

        bim_kpi_dict, trad_kpi_dict = self.load_kpi_values(grade)

        case = Case(
            case_id=f"{detailed_zoning} {grade}",
            kpi_grade=grade,
            kpi_bim=KPISet("BIM", bim_kpi_dict),
            kpi_trad=KPISet("TRAD", trad_kpi_dict),
            location_type=location_type,
            zoning=zoning,
            detailed_zoning=detailed_zoning,
            far=far,
            bcr=None
        )

        return case

    def _determine_grade_from_excel(self, location_type: str, zoning: str,
                                     detailed_zoning: str, far: float) -> str:
        """
        CASE 시트에서 등급을 찾아서 반환

        Args:
            location_type: 도심/외곽
            zoning: 상업지역, 주거지역 등
            detailed_zoning: 중심상업지역, 제1종 일반주거지역 등
            far: 용적률 (%)

        Returns:
            등급 (A, B, C, D)
        """
        ws = self.workbook['CASE']

        # Row 2에서 헤더 찾기 (도심/외곽)
        location_col = None
        for col_idx in range(4, ws.max_column + 1):
            header = ws.cell(2, col_idx).value
            if header and location_type in str(header):
                location_col = col_idx
                break

        if location_col is None:
            # 기본값: 도심이면 col 5, 외곽이면 col 6
            location_col = 5 if "도심" in location_type else 6

        # Row 11부터 데이터 시작, 매칭되는 행 찾기
        best_match_grade = "A"  # 기본값
        best_match_far = 0

        for row_idx in range(11, ws.max_row + 1):
            row_zoning = ws.cell(row_idx, 2).value
            row_far = self._safe_float(ws.cell(row_idx, 3).value)

            # 지역 매칭
            if row_zoning and zoning in str(row_zoning):
                # 용적률 매칭 (입력 FAR보다 작거나 같은 것 중 가장 큰 것)
                if row_far > 0 and row_far <= far and row_far > best_match_far:
                    # 해당 셀의 등급 추출
                    cell_value = ws.cell(row_idx, location_col).value
                    if cell_value:
                        # "중심상업지역 A" 같은 형식에서 등급 추출
                        grade = self._extract_grade(str(cell_value), detailed_zoning)
                        if grade:
                            best_match_grade = grade
                            best_match_far = row_far

        return best_match_grade

    def _extract_grade(self, cell_value: str, detailed_zoning: str) -> Optional[str]:
        """
        셀 값에서 등급 추출

        예: "중심상업지역 A" → "A"
        예: "중심상업지역 A\n일반상업지역 A" → "A" (detailed_zoning이 포함된 라인에서)
        """
        lines = cell_value.split('\n')

        for line in lines:
            # detailed_zoning이 포함되어 있으면 해당 라인에서 등급 추출
            if detailed_zoning in line:
                # 마지막 단어가 등급 (A, B, C, D)
                parts = line.strip().split()
                if parts and parts[-1] in ['A', 'B', 'C', 'D']:
                    return parts[-1]

        # detailed_zoning이 없으면 첫 번째 라인에서 등급 추출
        if lines:
            parts = lines[0].strip().split()
            if parts and parts[-1] in ['A', 'B', 'C', 'D']:
                return parts[-1]

        return None

    def load_issue_cards(self, family: str) -> List[IssueCard]:
        """
        이슈카드 로드

        Args:
            family: "BIM" 또는 "TRAD"

        Returns:
            IssueCard 리스트
        """
        if family == "BIM":
            sheet_name = '이슈 카드_BIM 사용 (90개)_유영민'
            kpi_keys = ["WD", "CD", "AF", "PL"]
        else:
            sheet_name = '이슈 카드_BIM 미사용 (90개)_유영민'
            kpi_keys = ["RR", "SR", "CR", "FC"]

        ws = self.workbook[sheet_name]
        issues = []

        # Row 2부터 데이터 시작 (Row 1은 헤더)
        for row_idx in range(2, ws.max_row + 1):
            issue_id = ws.cell(row_idx, 1).value
            if not issue_id:
                continue  # 빈 행 스킵

            # 기본 정보
            name = ws.cell(row_idx, 2).value or ""
            category = ws.cell(row_idx, 3).value or ""
            severity = ws.cell(row_idx, 4).value or "S2"
            phase = ws.cell(row_idx, 5).value or ""

            # 지연/비용
            delay_min = self._safe_float(ws.cell(row_idx, 6).value)
            delay_max = self._safe_float(ws.cell(row_idx, 7).value)
            cost_min = self._safe_float(ws.cell(row_idx, 8).value)
            cost_max = self._safe_float(ws.cell(row_idx, 9).value)

            # 상세설명
            description = ws.cell(row_idx, 10).value or ""

            # KPI 가중치 (Col 12-15)
            kpi_weights = {
                kpi_keys[0]: self._safe_float(ws.cell(row_idx, 12).value),
                kpi_keys[1]: self._safe_float(ws.cell(row_idx, 13).value),
                kpi_keys[2]: self._safe_float(ws.cell(row_idx, 14).value),
                kpi_keys[3]: self._safe_float(ws.cell(row_idx, 15).value),
            }

            # 공정률 (Col 17)
            progress_segment = str(ws.cell(row_idx, 17).value or "0-100")

            issue = IssueCard(
                issue_id=issue_id,
                name=name,
                category=category,
                severity=severity,
                phase=phase,
                delay_min_weeks=delay_min,
                delay_max_weeks=delay_max,
                cost_min_pct=cost_min,
                cost_max_pct=cost_max,
                description=description,
                kpi_weights=kpi_weights,
                progress_segment=progress_segment,
                family=family
            )
            issues.append(issue)

        return issues

    def _safe_float(self, value) -> float:
        """안전하게 float로 변환"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0


# 싱글톤 인스턴스
_loader_instance: Optional[ExcelDataLoader] = None


def get_loader() -> ExcelDataLoader:
    """로더 싱글톤 가져오기"""
    global _loader_instance
    if _loader_instance is None:
        _loader_instance = ExcelDataLoader()
    return _loader_instance
